import csv
import io
import zipfile
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


class ImportFileError(ValueError):
    pass


@dataclass
class ParsedFile:
    file_format: str
    worksheet_names: list[str]
    selected_worksheet: str | None
    headers: list[str]
    rows: list[tuple[int, dict[str, Any], list[dict[str, str]]]]


def _safe_cell(value: Any, maximum: int) -> Any:
    if value is None:
        return None
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    if isinstance(value, bool):
        return "true" if value else "false"
    text = str(value).strip()
    if len(text) > maximum:
        raise ImportFileError("A cell exceeds the configured length limit")
    return text or None


def detect_format(filename: str, content_type: str | None, sample: bytes) -> str:
    suffix = Path(filename).suffix.lower()
    if suffix not in {".csv", ".xlsx"}:
        raise ImportFileError("Unsupported import format; use CSV or XLSX")
    if not sample:
        raise ImportFileError("The uploaded file is empty")
    if suffix == ".xlsx":
        if not sample.startswith(b"PK\x03\x04"):
            raise ImportFileError("File content does not match the XLSX extension")
        if content_type and content_type not in {"application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", "application/octet-stream"}:
            raise ImportFileError("File content type does not match the XLSX extension")
        return "xlsx"
    if sample.startswith((b"PK\x03\x04", b"MZ", b"%PDF", b"\x89PNG", b"GIF8")) or b"\x00" in sample:
        raise ImportFileError("File content does not match the CSV extension")
    if content_type and content_type not in {"text/csv", "text/plain", "application/csv", "application/vnd.ms-excel", "application/octet-stream"}:
        raise ImportFileError("File content type does not match the CSV extension")
    return "csv"


def _csv_file(path: Path, limits: dict) -> ParsedFile:
    try:
        text = path.read_text(encoding="utf-8-sig")
    except UnicodeDecodeError as exc:
        raise ImportFileError("CSV must use UTF-8 encoding") from exc
    if not text.strip():
        raise ImportFileError("The uploaded file is empty")
    sample = text[:8192]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;")
    except csv.Error:
        dialect = csv.excel
    csv.field_size_limit(limits["maximum_cell_length"])
    try:
        records = list(csv.reader(io.StringIO(text, newline=""), dialect=dialect, strict=True))
    except (csv.Error, OverflowError) as exc:
        raise ImportFileError("CSV structure is malformed") from exc
    header_index = next((index for index, row in enumerate(records[:20]) if any(str(cell).strip() for cell in row)), None)
    if header_index is None:
        raise ImportFileError("A meaningful header row was not found")
    headers = [str(value).strip() for value in records[header_index]]
    if not any(headers) or len(headers) > limits["maximum_columns"] or any(not header for header in headers):
        raise ImportFileError("CSV header is empty or invalid")
    rows = []
    for source_row, values in enumerate(records[header_index + 1:], start=header_index + 2):
        if not any(str(value).strip() for value in values):
            continue
        warnings = []
        if len(values) != len(headers):
            warnings.append({"field": "row", "code": "column_count", "message": "Row column count differs from the header"})
        padded = (values + [""] * len(headers))[:len(headers)]
        rows.append((source_row, {header: _safe_cell(value, limits["maximum_cell_length"]) for header, value in zip(headers, padded)}, warnings))
        if len(rows) > limits["maximum_rows"]:
            raise ImportFileError("File exceeds the configured row limit")
    if not rows:
        raise ImportFileError("The file contains no data rows")
    return ParsedFile("csv", [], None, headers, rows)


def _xlsx_file(path: Path, limits: dict, worksheet: str | None) -> ParsedFile:
    try:
        with zipfile.ZipFile(path) as archive:
            names = set(archive.namelist())
            if "[Content_Types].xml" not in names or "xl/workbook.xml" not in names:
                raise ImportFileError("File content is not a valid XLSX workbook")
            if any(name.lower().endswith("vbaproject.bin") for name in names):
                raise ImportFileError("Macro-enabled workbooks are not supported")
            if len(names) > 5000 or sum(item.file_size for item in archive.infolist()) > 100 * 1024 * 1024:
                raise ImportFileError("Workbook expanded content exceeds the safety limit")
        workbook = load_workbook(path, read_only=True, data_only=False, keep_links=False)
    except ImportFileError:
        raise
    except Exception as exc:
        raise ImportFileError("Workbook is corrupt, unreadable, or password protected") from exc
    try:
        names = workbook.sheetnames
        if not names or len(names) > limits["maximum_worksheets"]:
            raise ImportFileError("Workbook has no usable worksheet or exceeds the worksheet limit")
        selected = worksheet
        if selected is None:
            for candidate in names:
                candidate_rows = workbook[candidate].iter_rows()
                if any(any(cell.value is not None and str(cell.value).strip() for cell in row) for _, row in zip(range(20), candidate_rows)):
                    selected = candidate
                    break
        if selected is None:
            raise ImportFileError("Workbook has no suitable worksheet")
        if selected not in names:
            raise ImportFileError("Selected worksheet does not exist")
        sheet = workbook[selected]
        iterator = sheet.iter_rows()
        header_cells = None
        header_row = 0
        for row_number, cells in enumerate(iterator, start=1):
            if any(cell.value is not None and str(cell.value).strip() for cell in cells):
                header_cells, header_row = cells, row_number
                break
            if row_number >= 20:
                break
        if header_cells is None:
            raise ImportFileError("A meaningful header row was not found")
        headers = [str(cell.value).strip() if cell.value is not None else "" for cell in header_cells]
        if not any(headers) or len(headers) > limits["maximum_columns"] or any(not header for header in headers):
            raise ImportFileError("Workbook header is empty or invalid")
        rows = []
        for source_row, cells in enumerate(iterator, start=header_row + 1):
            if not any(cell.value is not None and str(cell.value).strip() for cell in cells):
                continue
            warnings = []
            values = []
            for cell in cells[:len(headers)]:
                if cell.data_type == "f":
                    warnings.append({"field": headers[cell.column - 1], "code": "formula_ignored", "message": "Formula was not evaluated"})
                    values.append(None)
                else:
                    values.append(_safe_cell(cell.value, limits["maximum_cell_length"]))
            values += [None] * (len(headers) - len(values))
            rows.append((source_row, dict(zip(headers, values)), warnings))
            if len(rows) > limits["maximum_rows"]:
                raise ImportFileError("Workbook exceeds the configured row limit")
        if not rows:
            raise ImportFileError("The worksheet contains no data rows")
        return ParsedFile("xlsx", names, selected, headers, rows)
    finally:
        workbook.close()


def parse_file(path: Path, file_format: str, limits: dict, worksheet: str | None = None) -> ParsedFile:
    return _csv_file(path, limits) if file_format == "csv" else _xlsx_file(path, limits, worksheet)
